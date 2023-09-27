from bs4 import BeautifulSoup
import urllib.request as req
import openpyxl
import os
from openpyxl.drawing.image import Image



#엑셀 파일 존재하는지 확인
if not os.path.exists("./melonchart.xlsx"):
    book=openpyxl.Workbook()
    book.save("./melonchart.xlsx")
else:
    book=openpyxl.Workbook()

# code=req.urlopen("https://www.melon.com/chart/index.htm") 이렇게 하면 웹페이지가 접근을 차단해기 떄문에 오류가 발생한다. 따라서 아래 코드를 사용
url="https://www.melon.com/chart/index.htm"
headers=req.Request(url,headers={"User-Agent":"Mozilla/5.0"})
code=req.urlopen(headers)
soup=BeautifulSoup(code,"html.parser")
title=soup.select("div.ellipsis.rank01 a") #속성값에 띄어쓰기는 .으로 표시
name=soup.select("div.ellipsis.rank02>span.checkEllipsis")
albumtitle=soup.select("div.ellipsis.rank03>a")
img=soup.select("a.image_typeAll>img")



openpyxl.load_workbook("./melonchart.xlsx") #엑셀 로드
#sheet=book["Sheet1"] #엑셀 중 Sheet1 사용
sheet=book.active #자동으로 열리는 시트를 사용. 뒤에 () 없음에 주의
sheet.column_dimensions["A"].width="15" #셀 너비 조정

row_num=1

#이미지 저장할 폴더 만들기
if not os.path.exists("./멜론이미지_실습"):
    os.mkdir("./멜론이미지_실습") #디렉토리생성

sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 29
sheet.column_dimensions["D"].width = 47


for i in range(len(title)):
    req.urlretrieve(img[i].attrs["src"],"./melonImg/{}.png".format(row_num)) #이미지 다운로드 (값,디렉토리&파일명)
           
    print(title[i].string, name[i].text, albumtitle[i].string, img[i].attrs["src"]) 
    sheet.row_dimensions[row_num].height=95
    #html태그 중 src 속성명 안에 이미지 링크가 있음. 또한 요소 중 요소가 있는 경우(name의 경우 span 안에 a태그명을 가진 요소에 이름이 있음)
    #해결법- .string 써서 none 나오면 .text 쓰기
    imgExcel=Image("./melonImg/{}.png".format(row_num)) #순서 주의! 이미지가 1열, 다음부터가 2열임.
    sheet.add_image(imgExcel, "A{}".format(row_num)) #이미지 삽입 
    sheet.cell(row=row_num,column=2).value=title[i].string #value 전에는 해당 행과 열 클릭, value 뒤에는 엑셀에 값 쓰기
    sheet.cell(row=row_num,column=3).value=name[i].text
    sheet.cell(row=row_num,column=4).value=albumtitle[i].string
    
    
    
    book.save("/Users/admin/Documents/visual studio code/Python/멜론차트.xlsx") #저장
    row_num+=1
    
    

    
    
    