import openpyxl
from openpyxl.styles import * 
import pandas as pd
from openpyxl.chart import Reference, Series, LineChart, BarChart, PieChart

df=pd.read_excel("./엑셀데이터/아파트분양가격.xlsx")
locations=df["지역명"].unique() #["지역명"] 열의 있는 데이터 중복값을

for i in locations:
    writer=pd.ExcelWriter("./엑셀데이터/아파트분양가격_지역별/아파트분양가격_{}.xlsx".format(i))
    df_location=df[df["지역명"]==i] #지역별 아파트 데이터 저장
    years=df_location["연도"].unique() #지역별 데이터에서 연도만 추출
    for y in years:
        df_year=df_location[df_location["연도"]==y]
        df_result=df_year.sort_values(by="월",ascending=True) #오름차순 정렬
        df_result.to_excel(writer,sheet_name="{}년".format(y)) #pandas로 시트 이름 지정해서 쓰기
    writer.save()
    print("{}지역 엑셀파일 분류 완료".format(i))
    
    
#서식 만들기
myfont=Font(name="맑은 고딕",size=12,bold=True)
myAlign=Alignment(horizontal="center") #가운데 정렬
orangeColor=PatternFill(patternType="solid",fgColor=Color("FE9A2E")) #색 가져오기 (#제외!)
mintColor=PatternFill(patternType="solid",fgColor=Color("67F8E9"))
yellowColor=PatternFill(patternType="solid",fgColor=Color("FBFB57"))
myborder=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin")) #셀 테두리 지정
    
    
for i in locations:
    book=openpyxl.load_workbook("./엑셀데이터/아파트분양가격_지역별/아파트분양가격_{}.xlsx".format(i))
    # sheet=book.active 하면 처음 엑셀 열자마자 보이는 시트(첫 시트)사용하기 떄문에 아래처럼 지정해서 사용
    for sheetname in book.sheetnames: #book.sheetnames는 시트네임을 리스트로 갖고 있음
        sheet=book[sheetname] #시트 하나 가져오기
        for row in sheet["B1:F1"]: #B1->F1의 범위 (제한된 행 또는 열의 연속) 가져오기
            for cell in row: #각각의 셀 지정
                cell.font=myfont
                cell.alignment=myAlign
                cell.fill=orangeColor
                cell.border=myborder
        for row in sheet["B2:F{}".format(sheet.max_row)]: #데이터가 있는 최대 행 반환 (열은 알파벳, 행은 숫자임을 기억하자)
            for cell in row:
                cell.alignment=myAlign
                cell.fill=mintColor
                cell.border=myborder
                
        
        #차트 그리기
        chart=BarChart()
        chart.title="{}지역 {}년도 아파트 분양가격".format(i,sheetname)
        value=Reference(sheet,range_string="{}!F2:F{}".format(sheetname,sheet.max_row))
        value_series=Series(value,title="분양가격")  
        chart.append(value_series)
        sheet.add_chart(chart,"G1")
    book.save("./엑셀데이터/아파트분양가격_지역별/아파트분양가격_{}.xlsx".format(i))
    print("{}지역 엑셀 서식 지정 완료".format(i))

            
             
