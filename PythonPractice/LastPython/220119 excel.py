import openpyxl
#from openpyxl.styles import 어쩌구 하는 것 대신
from openpyxl.styles import * #모든 함수 가져옴

book=openpyxl.load_workbook("./엑셀데이터/2017광고비total.xlsx")
sheet=book.active #() 없음

#서식 만들기
myfont=Font(name="맑은 고딕",size=12,bold=True)
myAlign=Alignment(horizontal="center") #가운데 정렬
 
orangeColor=PatternFill(patternType="solid",fgColor=Color("FE9A2E")) #색 가져오기 (#제외!)
mintColor=PatternFill(patternType="solid",fgColor=Color("67F8E9"))
yellowColor=PatternFill(patternType="solid",fgColor=Color("FBFB57"))

myborder=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin")) #셀 테두리 지정


#sheet.call(row=14, column=1).value 대신
sheet["A14"].value="합계"
sheet["A14"].font=myfont
sheet["A14"].alignment=myAlign
sheet["A14"].fill=orangeColor
sheet["A14"].border=myborder

#여러 셀 서식지정 한번에
for row in sheet["B14:D14"]: #B14행부터 D14행까지 (사실상 B~D로 열만 달라지는 것)
    for cell in row:
        cell.font=myfont
        cell.alignment=myAlign
        cell.fill=orangeColor
        cell.border=myborder

for row in sheet["B2:D13"]: 
    for cell in row:
        cell.font=myfont
        cell.alignment=myAlign
        cell.fill=mintColor
        cell.border=myborder


sheet["B14"].value="SUM(B2:B13)" #엑셀 내장함수 사용
sheet["C14"].value="SUM(C2:C13)" 
sheet["D14"].value="SUM(D2:D13)" 

book.save("./엑셀데이터/2017광고비total(서식지정).xlsx")
